import openpyxl

# create blank wb
wb = openpyxl.Workbook()
print ( wb )

print(wb.sheetnames)

sheet = wb["Sheet"]
print ( sheet )

print(sheet["A1"].value)
print(sheet["A1"].value == None)

# edit cells
sheet["A1"] = 42
sheet["A2"] = "Hello"

# save sheet to hard disk
wb.save("example.xlsx")

# create sheet
sheet2 = wb.create_sheet()
print(wb.sheetnames)
print(sheet2.title)

# edit sheet name
sheet2.title = "mysheetname"
print(wb.sheetnames)

wb.save("example.xlsx")

# create sheet part 2
wb.create_sheet(index=0, title="asdsheet")
print(wb.sheetnames)
